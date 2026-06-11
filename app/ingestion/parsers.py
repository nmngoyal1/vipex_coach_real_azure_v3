from __future__ import annotations
import csv
import io
import json
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO
from openpyxl import load_workbook

SUPPORTED_TEXT_COLUMNS = {"raw_text", "idea", "ideas", "raw idea", "workshop idea", "text"}

@dataclass
class ParsedIdeas:
    raw_ideas: list[str]
    parser: str
    warnings: list[str]
    metadata: dict


def _normalise_lines(lines: list[str]) -> list[str]:
    out: list[str] = []
    for line in lines:
        value = str(line).strip()
        if value and value.lower() not in SUPPORTED_TEXT_COLUMNS:
            out.append(value)
    return out


def parse_txt(content: bytes) -> ParsedIdeas:
    text = content.decode("utf-8-sig", errors="replace")
    return ParsedIdeas(_normalise_lines(text.splitlines()), "txt", [], {"encoding": "utf-8-sig"})


def parse_csv(content: bytes) -> ParsedIdeas:
    text = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.DictReader(io.StringIO(text)))
    warnings: list[str] = []
    if rows and rows[0]:
        headers = {h.strip().lower(): h for h in rows[0].keys() if h}
        chosen = None
        for name in SUPPORTED_TEXT_COLUMNS:
            if name in headers:
                chosen = headers[name]
                break
        if chosen:
            return ParsedIdeas([r[chosen].strip() for r in rows if r.get(chosen, "").strip()], "csv", warnings, {"column": chosen})
        warnings.append("No raw_text/idea column found; falling back to first non-empty column.")
        first_col = next(iter(rows[0].keys()))
        return ParsedIdeas([r[first_col].strip() for r in rows if r.get(first_col, "").strip()], "csv", warnings, {"column": first_col})
    return parse_txt(content)


def parse_json(content: bytes) -> ParsedIdeas:
    data = json.loads(content.decode("utf-8-sig"))
    if isinstance(data, dict):
        for key in ["raw_ideas", "ideas", "items"]:
            if key in data and isinstance(data[key], list):
                data = data[key]
                break
    if not isinstance(data, list):
        raise ValueError("JSON upload must be a list or contain raw_ideas/ideas/items list")
    raw: list[str] = []
    for item in data:
        if isinstance(item, str):
            raw.append(item)
        elif isinstance(item, dict):
            for key in ["raw_text", "idea", "text"]:
                if item.get(key):
                    raw.append(str(item[key]))
                    break
    return ParsedIdeas(_normalise_lines(raw), "json", [], {})


def parse_xlsx(content: bytes) -> ParsedIdeas:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ParsedIdeas([], "xlsx", ["Workbook has no rows."], {"sheet": ws.title})
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    chosen_idx = None
    for idx, h in enumerate(header):
        if h in SUPPORTED_TEXT_COLUMNS:
            chosen_idx = idx
            break
    warnings: list[str] = []
    if chosen_idx is None:
        chosen_idx = 0
        warnings.append("No raw_text/idea header found; using first column.")
    values = []
    for row in rows[1:]:
        if len(row) > chosen_idx and row[chosen_idx] is not None:
            values.append(str(row[chosen_idx]))
    return ParsedIdeas(_normalise_lines(values), "xlsx", warnings, {"sheet": ws.title, "column_index": chosen_idx})


def parse_image_or_pdf_placeholder(content: bytes, filename: str) -> ParsedIdeas:
    # The case study says inputs can include whiteboard screenshots, handwritten sticky-notes, and PDFs.
    # If Azure AI Document Intelligence is configured, extract OCR lines. Otherwise return a clear warning.
    try:
        from app.settings import get_settings
        s = get_settings()
        if s.document_intelligence_endpoint and s.document_intelligence_key:
            from app.azure.document_intelligence_adapter import AzureDocumentIntelligenceExtractor
            lines = AzureDocumentIntelligenceExtractor().extract_lines(content)
            return ParsedIdeas(_normalise_lines(lines), "azure_document_intelligence_ocr", [], {"filename": filename, "bytes": len(content), "lines": len(lines)})
    except Exception as exc:
        return ParsedIdeas([], "ocr_failed", [f"OCR extraction failed for {filename}: {exc}"], {"filename": filename, "bytes": len(content)})
    return ParsedIdeas(
        [],
        "ocr_required",
        [f"{filename} requires OCR/document extraction. Configure Azure AI Document Intelligence or pre-convert to CSV/XLSX/TXT."],
        {"filename": filename, "bytes": len(content)},
    )


def extract_raw_ideas_from_upload(filename: str, content: bytes) -> ParsedIdeas:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md"}:
        return parse_txt(content)
    if suffix == ".csv":
        return parse_csv(content)
    if suffix == ".json":
        return parse_json(content)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(content)
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".pdf", ".tiff"}:
        return parse_image_or_pdf_placeholder(content, filename)
    raise ValueError(f"Unsupported upload type: {suffix}. Supported: csv, xlsx, json, txt; OCR seam for image/pdf.")
